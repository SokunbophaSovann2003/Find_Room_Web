"""Build the QA Testing Report Excel workbook for findroom_web."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

TESTER = "QA Bot (Claude)"
TEST_DATE = "2026-05-20"

# ---------- styling helpers -----------------------------------------------

THIN = Side(border_style="thin", color="CFD8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F6F4A")  # brand green
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SECTION_FILL = PatternFill("solid", fgColor="E8F5EE")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="0F3D2A")

STATUS_FILLS = {
    "Pass": PatternFill("solid", fgColor="DCFCE7"),
    "Failed": PatternFill("solid", fgColor="FEE2E2"),
    "Pending": PatternFill("solid", fgColor="FEF9C3"),
}
STATUS_FONTS = {
    "Pass": Font(name="Calibri", size=10, bold=True, color="14532D"),
    "Failed": Font(name="Calibri", size=10, bold=True, color="7F1D1D"),
    "Pending": Font(name="Calibri", size=10, bold=True, color="713F12"),
}

SEVERITY_FILLS = {
    "Critical": PatternFill("solid", fgColor="DC2626"),
    "High": PatternFill("solid", fgColor="F97316"),
    "Medium": PatternFill("solid", fgColor="FACC15"),
    "Low": PatternFill("solid", fgColor="93C5FD"),
    "": PatternFill("solid", fgColor="F1F5F9"),
}
SEVERITY_FONTS = {
    "Critical": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "High": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "Medium": Font(name="Calibri", size=10, bold=True, color="1F2937"),
    "Low": Font(name="Calibri", size=10, bold=True, color="1F2937"),
    "": Font(name="Calibri", size=10, color="475569"),
}

WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

COLUMNS = [
    ("Module/Page", 18),
    ("Feature", 22),
    ("Test Scenario", 34),
    ("Steps to Test", 42),
    ("Expected Result", 34),
    ("Actual Result", 34),
    ("Status", 11),
    ("Severity", 11),
    ("Screenshot / Reference", 26),
    ("Remark / Comment", 30),
    ("Tested By", 16),
    ("Testing Date", 13),
]

# ---------- test cases ----------------------------------------------------
# Status: Pass | Failed | Pending
# Severity: Critical | High | Medium | Low | "" (for Pass rows leave blank)

TEST_CASES = [
    # Module: Explore Page ------------------------------------------------
    {
        "module": "Explore Page",
        "feature": "Initial render",
        "scenario": "Open /explore and verify hero, search, room grid render",
        "steps": "1) Navigate to http://localhost:3000/\n2) Page should redirect to /explore\n3) Observe hero, search bar and listing grid",
        "expected": "Hero heading, sub-tagline, 3-input search bar, List/Map tabs and room cards visible. Counter shows 'N rooms available right now'.",
        "actual": "Hero, search bar (Where to / Property type / Sort by price + Search button), List/Map tabs and 29 mock listings rendered. Counter reads '29 rooms available right now'.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/01_explore_desktop.png",
        "remark": "Page redirect from / to /explore works.",
    },
    {
        "module": "Explore Page",
        "feature": "Location filter (province)",
        "scenario": "Cascading location picker filters listings",
        "steps": "1) Click 'Where to?' button\n2) Pick 'Phnom Penh' -> 'Chamkar Mon'\n3) Click 'Show all rooms in Chamkar Mon'",
        "expected": "Listing count updates to only listings in selected district; chip pill shows 'Phnom Penh, Chamkar Mon' with clear (×) affordance.",
        "actual": "Counter dropped from 29 to 6 rooms. Selected chip with × button rendered correctly.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/02_explore_filter_chamkar_mon.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Property type filter",
        "scenario": "Selecting a property type narrows results",
        "steps": "1) Open 'Any Property Type' dropdown\n2) Pick 'Room'",
        "expected": "Counter reflects rooms-only count; pluralisation correct.",
        "actual": "After picking Room (with Chamkar Mon still set) counter read '1 room available right now'. Singular form correct.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/03_explore_filter_room.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Sort by price",
        "scenario": "Sort order applies to current filtered set",
        "steps": "1) Open 'Sort by price' dropdown\n2) Select 'Price: low to high'",
        "expected": "Cards reorder by price ascending; selected label persists on the button.",
        "actual": "Sort options dropdown opened with three choices, low-to-high applied and label updated on the button.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/04_explore_sort.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Clear location chip",
        "scenario": "× on selected chip removes that filter",
        "steps": "1) With Phnom Penh / Chamkar Mon set, click × on the chip",
        "expected": "Location filter cleared, other filters (type, sort) remain.",
        "actual": "Location cleared, counter returned to 3 rooms (Room type still active). Other filters preserved as expected.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/05_explore_clear_location.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Map view tab",
        "scenario": "Switch from list to map view",
        "steps": "1) Click 'Map' tab",
        "expected": "List replaced by interactive Leaflet map; counter switches phrasing to 'N rooms in this area'.",
        "actual": "Leaflet map rendered with tiles, counter changed to '1 room in this area'.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/06_explore_map.png",
        "remark": "Counter phrasing differs (in this area vs available right now); this is by design but worth a doc note.",
    },

    # Module: Room Detail -------------------------------------------------
    {
        "module": "Room Detail",
        "feature": "Page render",
        "scenario": "/rooms/2 displays hero, stats, description, fees and host card",
        "steps": "1) From explore click a room card\n2) Verify hero image, title, address, $price, bedrooms/area/floor stats, description, amenities and fees table",
        "expected": "All sections render; Back link visible; sidebar shows host card and location card.",
        "actual": "Hero image with property type badge, title 'Bright shared room in BKK1', address, price $180/month, stats (1 Bed / 18m² / 2 Floor), 'About this place', 'What this place offers' (4 amenities), Fees & utilities table (rent / deposit / electricity / water) and host card with phone + Telegram all rendered.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/07_room_detail.png",
        "remark": "",
    },
    {
        "module": "Room Detail",
        "feature": "Image gallery (lightbox)",
        "scenario": "'3 photos' button opens full-screen viewer",
        "steps": "1) Click the '3 photos' button on the hero",
        "expected": "Full-screen viewer covers viewport, shows all photos with thumbnails; Esc / close button dismisses.",
        "actual": "Lightbox opens, but photos render as a tiled grid that does NOT fill the viewport — large empty area visible and the page header bleeds through behind the overlay.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "screenshots_fixed/03_room_detail.png",
        "remark": "FIXED 2026-05-20 — rewrote viewer in src/components/ImageGallery.tsx as a single-image carousel with prev/next chevrons + thumbnail strip; backdrop changed from bg-ink/95 to bg-ink (fully opaque). Retested: lightbox now covers viewport completely.",
    },
    {
        "module": "Room Detail",
        "feature": "Phone & Telegram links",
        "scenario": "Host contact links open native handlers",
        "steps": "1) Inspect <a href> on host card",
        "expected": "Phone link uses tel:+855... and Telegram link uses https://t.me/+855...",
        "actual": "Phone href = tel:+85517888112; Telegram href = https://t.me/+85517888112. Both formats correct.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/07_room_detail.png",
        "remark": "",
    },
    {
        "module": "Room Detail",
        "feature": "Deferred Google Maps embed",
        "scenario": "'Show map' lazy-loads Google Maps iframe",
        "steps": "1) Scroll to Location card\n2) Click 'Show map'",
        "expected": "Iframe loads with src pointing at google.com/maps?q=lat,lng",
        "actual": "Iframe injected with src=https://www.google.com/maps?q=11.5444%2C104.9263&z=15&output=embed",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/09_room_show_map.png",
        "remark": "Lazy load avoids loading Maps on initial paint — good for performance.",
    },
    {
        "module": "Room Detail",
        "feature": "Missing room fallback",
        "scenario": "Visit /rooms/9999 (non-existent id)",
        "steps": "1) Navigate to /rooms/9999",
        "expected": "Friendly empty state with 'Back to Explore' CTA.",
        "actual": "'Room not found' heading, supporting copy and 'Back to Explore' button rendered correctly.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/10_room_not_found.png",
        "remark": "",
    },
    {
        "module": "Room Detail",
        "feature": "Similar rooms",
        "scenario": "Bottom section shows up to 3 similar listings",
        "steps": "1) Scroll to the bottom of /rooms/2",
        "expected": "3 'Similar rooms' cards each linking to its own detail page.",
        "actual": "3 cards rendered (Budget room for students near RUPP, Backpacker room near Pub Street, Cosy room in Tuol Tom Poung (Occupied)).",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/07_room_detail.png",
        "remark": "",
    },

    # Module: Authentication ---------------------------------------------
    {
        "module": "Authentication",
        "feature": "AuthGuard redirect",
        "scenario": "Hitting /profile while logged out opens AuthModal",
        "steps": "1) Clear localStorage 'findroom.session'\n2) Navigate to /profile",
        "expected": "AuthGuard mounts AuthModal in Login mode; underlying page is gated behind backdrop.",
        "actual": "AuthModal opened with 'Welcome back' heading, phone+password fields and 'No account yet? Create one' link.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/11_auth_modal_login.png",
        "remark": "",
    },
    {
        "module": "Authentication",
        "feature": "Login — empty submit",
        "scenario": "Submitting the login form with empty fields",
        "steps": "1) Click 'Log in' with no values entered",
        "expected": "Native HTML5 'required' validation prevents submit and focuses the first empty field.",
        "actual": "Browser focused the phone input; no request sent.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/11_auth_modal_login.png",
        "remark": "Relies on native validation only — fine for current scope.",
    },
    {
        "module": "Authentication",
        "feature": "Login — phone format",
        "scenario": "Enter a phone with fewer than 8 digits",
        "steps": "1) Type '1234' in phone\n2) Type 'somepass' in password\n3) Click 'Log in'",
        "expected": "Inline error: 'Enter a valid Cambodian phone number (8-9 digits after +855).'",
        "actual": "Exact error message shown in red inline banner; submit blocked.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/12_auth_invalid_phone.png",
        "remark": "",
    },
    {
        "module": "Authentication",
        "feature": "Register — switch to signup",
        "scenario": "Click 'Create one' to switch tabs",
        "steps": "1) On login modal click 'Create one'",
        "expected": "Modal switches to 'Create your account' with name + phone + password fields and 'At least 8 characters' password hint.",
        "actual": "Modal swapped tabs as expected; placeholder hint visible.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/13_auth_register.png",
        "remark": "",
    },
    {
        "module": "Authentication",
        "feature": "Register — short password",
        "scenario": "Submit register with a 3-character password",
        "steps": "1) Name=QA Tester\n2) Phone=87654321\n3) Password=123\n4) Click 'Create account'",
        "expected": "Inline validation: 'Password must be at least 8 characters.'",
        "actual": "Exact error displayed; account not created.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/14_auth_short_password.png",
        "remark": "",
    },
    {
        "module": "Authentication",
        "feature": "Register — successful signup",
        "scenario": "Submit register with valid credentials in demo mode",
        "steps": "1) Name=QA Tester\n2) Phone=87654321\n3) Password=password123\n4) Click 'Create account'",
        "expected": "Session created, modal closes, user lands on /profile with their entered username.",
        "actual": "Session stored (uid demo-85587654321, username 'QA Tester'); however the profile header rendered the username from a previous demo user ('FindRoom user') with that user's avatar — see profile-overrides bug below.",
        "status": "Pass",
        "severity": "High",
        "screenshot": "screenshots/15_profile_after_register.png",
        "remark": "FIXED 2026-05-20 — see linked fix below. Retested: new accounts now land on /profile with their own entered username and a clean (initial-only) avatar.",
    },
    {
        "module": "Authentication",
        "feature": "Profile overrides leak across users",
        "scenario": "A previous user's avatar and display name follow any new account on the same device",
        "steps": "1) Log in as user A and Edit Profile (change name + upload avatar)\n2) Log out\n3) Register or log in as user B\n4) Visit /profile",
        "expected": "Each session should see ONLY their own profile name and avatar.",
        "actual": "localStorage 'findroom.profile-overrides' is stored under a single global key (not keyed by uid). Confirmed user B sees user A's saved username 'FindRoom user' and base64 cat avatar.",
        "status": "Pass",
        "severity": "High",
        "screenshot": "screenshots/16_overrides_leak.png",
        "remark": "FIXED 2026-05-20 — src/lib/profile-overrides.ts now namespaces storage as findroom.profile-overrides.<uid>; legacy global key is purged on first load. Callers in profile/page.tsx, profile/list-room/page.tsx, components/Navbar.tsx updated to pass session.uid. Retested with two demo accounts: User B no longer inherits User A's name/avatar.",
    },
    {
        "module": "Authentication",
        "feature": "Log out",
        "scenario": "Log out from profile",
        "steps": "1) On /profile click 'Log out'",
        "expected": "Session cleared, redirect to /explore, navbar shows 'Log in' instead of avatar.",
        "actual": "Behaviour matches expectation (verified during session cleanup steps).",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "",
    },

    # Module: Profile Page ------------------------------------------------
    {
        "module": "Profile Page",
        "feature": "Empty listings state",
        "scenario": "Fresh user with no listings",
        "steps": "1) Clear local-rooms\n2) Reload /profile",
        "expected": "'No listings yet' card with 'Create your first listing' CTA.",
        "actual": "Empty state card with house icon, copy and CTA button shown.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/17_profile_empty.png",
        "remark": "",
    },
    {
        "module": "Profile Page",
        "feature": "Edit profile — save name",
        "scenario": "Update display name via Edit profile modal",
        "steps": "1) Click 'Edit profile'\n2) Change Name to 'QA Tester Updated'\n3) Click Save",
        "expected": "Modal closes, header shows new name immediately.",
        "actual": "Header updated to 'QA Tester Updated' without page reload.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/18_profile_edit_save.png",
        "remark": "",
    },
    {
        "module": "Profile Page",
        "feature": "Edit profile — modal copy",
        "scenario": "Edit profile modal copy & layout",
        "steps": "1) Open Edit profile modal",
        "expected": "Modal contains Avatar uploader, Name (required), Login phone (required) with helper text about contacts.",
        "actual": "All three controls present; helper text 'You sign in with this number. Changing it doesn't affect the contact info you publish on listings.' is helpful and clear.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/19_profile_edit_modal.png",
        "remark": "",
    },
    {
        "module": "Profile Page",
        "feature": "Listing action menu — Copy",
        "scenario": "Per-listing action menu duplicates a listing",
        "steps": "1) Click 3-dot 'Listing options' on a listing card\n2) Click 'Copy'",
        "expected": "Navigates to /profile/list-room?type=<type>&copyFrom=<id> with form pre-populated.",
        "actual": "URL = /profile/list-room?type=apartment&copyFrom=local-1779243463762. Title field auto-filled with 'Copy of QA Test Apartment', description and other fields carried over.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/20_listing_copy.png",
        "remark": "Photos correctly NOT copied (intended).",
    },
    {
        "module": "Profile Page",
        "feature": "Listing action menu — Mark as occupied / Delete",
        "scenario": "Menu exposes both actions",
        "steps": "1) Open per-listing 3-dot menu",
        "expected": "Items 'Mark as occupied', 'Copy', 'Delete' visible.",
        "actual": "All three items present in the floating menu (verified via DOM inspection).",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "End-to-end click of Mark/Delete not fully exercised due to menu being portalled outside the main tree — manual retest recommended.",
    },

    # Module: List Room Form ---------------------------------------------
    {
        "module": "List Room Form",
        "feature": "Property type picker",
        "scenario": "Picker presents 6 property types when entering the form via CTA",
        "steps": "1) From profile, click 'List a room' / 'Create your first listing'",
        "expected": "Modal lists Room, Apartment, Condo, Flat, House, Villa each with a one-line description.",
        "actual": "All 6 options shown with subtitles. Clicking 'Apartment' navigates to /profile/list-room?type=apartment.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/21_listroom_typepicker.png",
        "remark": "",
    },
    {
        "module": "List Room Form",
        "feature": "Required field — Title",
        "scenario": "Submit the form with no title",
        "steps": "1) Open List Room form\n2) Leave title empty\n3) Click 'Publish listing'",
        "expected": "Validation error appears under Title and form does not submit.",
        "actual": "Inline message 'Add a title for your listing.' shown; submit blocked.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/22_listroom_title_required.png",
        "remark": "",
    },
    {
        "module": "List Room Form",
        "feature": "Required field — Province",
        "scenario": "Submit with title + description but no location",
        "steps": "1) Fill title + description\n2) Click 'Publish listing'",
        "expected": "Error: 'Pick a province for your listing.'",
        "actual": "Exact message displayed.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/23_listroom_province_required.png",
        "remark": "",
    },
    {
        "module": "List Room Form",
        "feature": "Required field — Rent",
        "scenario": "Submit with title, description and province but no rent",
        "steps": "1) Set title, description, province=Phnom Penh / Chamkar Mon\n2) Click 'Publish listing'",
        "expected": "Error: 'Set a monthly rent amount.'",
        "actual": "Exact message displayed.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/24_listroom_rent_required.png",
        "remark": "",
    },
    {
        "module": "List Room Form",
        "feature": "Successful publish",
        "scenario": "Submit fully valid listing",
        "steps": "1) Title='QA Test Apartment'\n2) Description filled\n3) Province=Phnom Penh / district=Chamkar Mon\n4) Add fee, set Rent=$250 Monthly\n5) Click 'Publish listing'",
        "expected": "Form submits, user returns to /profile and new card appears under 'My listings'.",
        "actual": "Redirected to /profile. 'QA Test Apartment' card with $250 / month, Chamkar Mon, Phnom Penh shown. Counter 'My listings 1'.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/25_listroom_publish_success.png",
        "remark": "Listing also appears in /explore feed (merged with mock data).",
    },
    {
        "module": "List Room Form",
        "feature": "LocationPicker reused on listing form",
        "scenario": "'Show all rooms in X' link appears inside the listing creation flow",
        "steps": "1) Open List Room form\n2) Click 'Pick province'\n3) Pick a province and district",
        "expected": "Copy in the picker should match the context (e.g. 'Use this district' / 'Set district') rather than 'Show all rooms in X', which only makes sense on Explore.",
        "actual": "Top of the picker still reads 'Show all rooms in Chamkar Mon'. Users on the listing form may be confused or click it expecting to back out.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "screenshots/26_listroom_picker_copy.png",
        "remark": "FIXED 2026-05-20 — added `intent` prop to LocationPicker ('browse' default | 'select'); list-room form passes intent='select'. Picker now reads 'Use Phnom Penh' / 'Use Chamkar Mon' / 'Clear' on the listing form. Retested.",
    },
    {
        "module": "List Room Form",
        "feature": "Mobile sticky footer",
        "scenario": "Mobile viewport shows Clear / Publish at the bottom",
        "steps": "1) Resize to 375x812 (mobile)\n2) Open List Room form",
        "expected": "Sticky footer with 'Clear' and 'Publish' buttons; BottomNav hidden on this route.",
        "actual": "Sticky footer present, BottomNav correctly hidden on /profile/list-room.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/27_listroom_mobile_footer.png",
        "remark": "",
    },

    # Module: Global Nav --------------------------------------------------
    {
        "module": "Global Nav",
        "feature": "Navbar — desktop",
        "scenario": "Desktop navbar exposes List room, language and avatar links",
        "steps": "1) Resize to 1280x820\n2) Load /explore",
        "expected": "Logo (left), 'List room' button, language flag toggle, and avatar/login slot (right).",
        "actual": "Logo, 'List room' button, Cambodia flag toggle and 'Q' avatar shown.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/28_navbar_desktop.png",
        "remark": "",
    },
    {
        "module": "Global Nav",
        "feature": "BottomNav — mobile",
        "scenario": "BottomNav visible on listing/profile routes on mobile",
        "steps": "1) Resize to mobile preset\n2) Visit /explore",
        "expected": "Sticky bottom bar with Home, raised + FAB, Profile; safe-area padding for notched devices.",
        "actual": "BottomNav with Home / + FAB / Profile rendered. Tapping + on a logged-out session triggers AuthModal (verified earlier).",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/29_bottomnav_mobile.png",
        "remark": "Correctly hidden on /profile/list-room and /rooms/[id].",
    },
    {
        "module": "Global Nav",
        "feature": "Language toggle persistence",
        "scenario": "Toggle persists selected language across reloads",
        "steps": "1) Click language flag\n2) Reload\n3) Inspect localStorage 'findroom.language' and aria-label",
        "expected": "Selection persists; aria-label and flag swap accordingly.",
        "actual": "localStorage value flipped km <-> en; aria-label updated to 'Current language: English. Tap to switch to Khmer.' after reload.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/30_lang_toggle.png",
        "remark": "",
    },
    {
        "module": "Global Nav",
        "feature": "Language toggle — no UI translation",
        "scenario": "Switching language should translate UI strings",
        "steps": "1) Click language toggle to switch to km / en\n2) Reload /explore and inspect headings",
        "expected": "Visible copy (hero heading, search placeholders, etc.) should appear in the selected language.",
        "actual": "All visible copy remained in English regardless of preference (km vs en). The toggle is a no-op for end users; only the stored preference changes.",
        "status": "Pass",
        "severity": "Medium",
        "screenshot": "screenshots_fixed/01_explore_km_desktop.png",
        "remark": "FIXED 2026-05-20 — added in-file DICT + useT() hook in src/lib/language.ts; useLanguage now broadcasts a 'findroom:language-change' event so every consumer re-renders on toggle. New ExploreHero client component + SearchBar + ExploreRooms now translate hero, placeholders, type/sort labels, list/map tabs, and the rooms counter. Khmer translations need a native-speaker review before launch.",
    },
    {
        "module": "Global Nav",
        "feature": "Footer",
        "scenario": "Footer renders contact + meta info",
        "steps": "1) Scroll to footer on any page",
        "expected": "Logo + blurb, 'GET IN TOUCH' (+855 12 345 678, Telegram, hello@findroom.kh, Facebook), copyright, 'Made in Cambodia'.",
        "actual": "All four contact links + tagline + © 2026 FindRoom.KH + 'Made in Cambodia' visible.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/32_footer.png",
        "remark": "",
    },

    # Module: Responsive --------------------------------------------------
    {
        "module": "Responsive Design",
        "feature": "Mobile (375x812)",
        "scenario": "Smoke check key pages at mobile width",
        "steps": "1) Resize to mobile preset\n2) Browse /, /explore, /profile, /profile/list-room",
        "expected": "Single-column layouts, BottomNav visible where appropriate, no horizontal scroll, sticky CTAs accessible.",
        "actual": "Layouts collapse cleanly; BottomNav present on /explore and /profile; List Room sticky footer in place; no horizontal scroll detected.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/33_responsive_mobile.png",
        "remark": "",
    },
    {
        "module": "Responsive Design",
        "feature": "Tablet (768x1024)",
        "scenario": "Smoke check at tablet width",
        "steps": "1) Resize to tablet preset\n2) Load /explore",
        "expected": "Full desktop navbar visible, room grid 3 columns, search bar inputs readable.",
        "actual": "Navbar full; grid 3 columns. However, search bar input labels truncate ('Where to? Pr…', 'Any Property …') because the bar width is tight.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "screenshots_fixed/02_explore_tablet_stacked.png",
        "remark": "FIXED 2026-05-20 — switched SearchBar responsive breakpoints from sm: to lg: so the bar stacks vertically below 1024px. Each input gets full width on tablet and mobile, no truncation. Retested at 768x1024.",
    },
    {
        "module": "Responsive Design",
        "feature": "Desktop (1280x820)",
        "scenario": "Smoke check at desktop width",
        "steps": "1) Resize to desktop preset\n2) Load /explore",
        "expected": "Hero, full search bar, 4-column room grid, List/Map tabs aligned right.",
        "actual": "All elements rendered as expected.",
        "status": "Pass",
        "severity": "",
        "screenshot": "screenshots/01_explore_desktop.png",
        "remark": "",
    },
    {
        "module": "Responsive Design",
        "feature": "Mock data inconsistency — BKK1 in Chamkar Mon",
        "scenario": "Listing 'Bright shared room in BKK1' assigns area BKK1 to Chamkar Mon district",
        "steps": "1) Open /rooms/2 (or similar)\n2) Note address 'St. 294, BKK1, Chamkar Mon, Phnom Penh'",
        "expected": "BKK1 (Boeng Keng Kang 1) is currently part of the Boeng Keng Kang khan since the 2018 redistribution.",
        "actual": "Mock data treats BKK1 as part of Chamkar Mon. May be intentional shorthand but the LocationPicker district list under Chamkar Mon does not include BKK1.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "",
        "remark": "FIXED 2026-05-20 — moved room id=2 in src/lib/mock-data.ts and the matching seed in src/lib/local-rooms.ts from district='Chamkar Mon' to district='Boeng Keng Kang' (area='Boeng Keng Kang Muoy'), matching the canonical structure in src/lib/locations.ts. Retested: filtering by Boeng Keng Kang now returns 'Bright shared room in BKK1' (was previously hidden under Chamkar Mon); filtering by Chamkar Mon no longer mis-attributes it.",
    },
]


# ---------- workbook generation -------------------------------------------

def write_header_row(ws, row, columns):
    for idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 28


def write_case_row(ws, row, case):
    values = [
        case["module"],
        case["feature"],
        case["scenario"],
        case["steps"],
        case["expected"],
        case["actual"],
        case["status"],
        case["severity"],
        case["screenshot"],
        case["remark"],
        TESTER,
        TEST_DATE,
    ]
    row_failed = case["status"] == "Failed"
    for idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=idx, value=value)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=10)
        if row_failed:
            cell.fill = PatternFill("solid", fgColor="FFF1F2")

    status_cell = ws.cell(row=row, column=7)
    status_cell.fill = STATUS_FILLS[case["status"]]
    status_cell.font = STATUS_FONTS[case["status"]]
    status_cell.alignment = CENTER

    severity_key = case["severity"] if case["severity"] in SEVERITY_FILLS else ""
    severity_cell = ws.cell(row=row, column=8)
    severity_cell.fill = SEVERITY_FILLS[severity_key]
    severity_cell.font = SEVERITY_FONTS[severity_key]
    severity_cell.alignment = CENTER

    # rough auto-height: assume 60 chars per wrapped line, max 6 lines factor
    longest = max(len(str(v)) for v in values if v)
    approx_lines = max(2, min(8, longest // 50 + 1))
    ws.row_dimensions[row].height = 15 * approx_lines


def build_module_sheet(wb, module_name, cases):
    ws = wb.create_sheet(module_name[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    write_header_row(ws, 1, COLUMNS)
    for offset, case in enumerate(cases):
        write_case_row(ws, 2 + offset, case)


def build_all_sheet(wb):
    ws = wb.create_sheet("All Cases")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    write_header_row(ws, 1, COLUMNS)
    for offset, case in enumerate(TEST_CASES):
        write_case_row(ws, 2 + offset, case)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{1 + len(TEST_CASES)}"


def build_summary_sheet(wb):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "FindRoom.KH — QA Testing Report (Retested)"
    ws["B2"].font = Font(name="Calibri", size=20, bold=True, color="0F3D2A")
    ws.row_dimensions[2].height = 30

    ws["B3"] = "Build under test: src branch 'main' @ commit 34c77c4 + 2026-05-20 fix pass"
    ws["B3"].font = Font(name="Calibri", size=11, color="334155")

    ws["B4"] = f"Tester: {TESTER}   |   First test: {TEST_DATE}   |   Retest: {TEST_DATE}   |   Environment: localhost:3000 (next dev) + next build verified"
    ws["B4"].font = Font(name="Calibri", size=11, color="334155")

    # totals
    total = len(TEST_CASES)
    passed = sum(1 for c in TEST_CASES if c["status"] == "Pass")
    failed = sum(1 for c in TEST_CASES if c["status"] == "Failed")
    pending = sum(1 for c in TEST_CASES if c["status"] == "Pending")
    sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for c in TEST_CASES:
        if c["status"] == "Failed" and c["severity"] in sev_counts:
            sev_counts[c["severity"]] += 1

    headers = [
        ("Total cases", total, "1F6F4A"),
        ("Passed", passed, "16A34A"),
        ("Failed", failed, "DC2626"),
        ("Pending", pending, "CA8A04"),
    ]
    for idx, (label, value, color) in enumerate(headers):
        col = 2 + idx
        ws.cell(row=6, column=col, value=label).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws.cell(row=6, column=col).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=6, column=col).alignment = CENTER
        ws.cell(row=7, column=col, value=value).font = Font(name="Calibri", size=20, bold=True, color="1F2937")
        ws.cell(row=7, column=col).alignment = CENTER

    ws.cell(row=9, column=2, value="Failed cases by severity").font = Font(name="Calibri", size=12, bold=True, color="0F3D2A")
    sev_row = 10
    for idx, (sev, count) in enumerate(sev_counts.items()):
        col = 2 + idx
        ws.cell(row=sev_row, column=col, value=sev).font = SEVERITY_FONTS[sev]
        ws.cell(row=sev_row, column=col).fill = SEVERITY_FILLS[sev]
        ws.cell(row=sev_row, column=col).alignment = CENTER
        ws.cell(row=sev_row + 1, column=col, value=count).font = Font(name="Calibri", size=16, bold=True)
        ws.cell(row=sev_row + 1, column=col).alignment = CENTER

    ws.cell(row=13, column=2, value="Cases fixed in this retest").font = Font(name="Calibri", size=12, bold=True, color="0F3D2A")
    # After the 2026-05-20 fix pass, the issues we want highlighted are the
    # ones that flipped from Failed → Pass (status now "Pass" + a "FIXED " note
    # in the remark column). Sort by their original severity.
    top_issues = [c for c in TEST_CASES if c["status"] == "Pass" and c["remark"].startswith("FIXED")]
    top_issues.sort(key=lambda c: {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}.get(c["severity"], 4))

    issue_header_row = 14
    issue_columns = [
        ("Severity", 12),
        ("Module", 18),
        ("Feature", 24),
        ("Summary", 60),
        ("Reference", 24),
    ]
    for idx, (name, width) in enumerate(issue_columns, start=2):
        cell = ws.cell(row=issue_header_row, column=idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(idx)].width = width

    for offset, c in enumerate(top_issues):
        row = issue_header_row + 1 + offset
        sev = c["severity"] if c["severity"] in SEVERITY_FILLS else ""
        sev_cell = ws.cell(row=row, column=2, value=c["severity"])
        sev_cell.fill = SEVERITY_FILLS[sev]
        sev_cell.font = SEVERITY_FONTS[sev]
        sev_cell.alignment = CENTER
        for idx, value in enumerate([c["module"], c["feature"], c["actual"], c["screenshot"]], start=3):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.alignment = WRAP
            cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[row].height = 60

    ws.column_dimensions["A"].width = 3


def main():
    wb = Workbook()
    wb.remove(wb.active)  # we'll add our own first

    # Group by module preserving original order
    modules = []
    seen = set()
    for c in TEST_CASES:
        if c["module"] not in seen:
            seen.add(c["module"])
            modules.append(c["module"])

    for module in modules:
        cases = [c for c in TEST_CASES if c["module"] == module]
        build_module_sheet(wb, module, cases)

    build_all_sheet(wb)
    build_summary_sheet(wb)

    out = "/Users/kosign/Documents/Startup/Find Room/findroom_web/qa_report/QA_Report_FindRoom_2026-05-20.xlsx"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
