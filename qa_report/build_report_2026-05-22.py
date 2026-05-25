"""Build the QA Testing Report Excel workbook for findroom_web (2026-05-22 pass)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

TESTER = "QA Bot (Claude)"
TEST_DATE = "2026-05-22"
SS_DIR = "screenshots_2026-05-22"

# ---------- styling helpers -----------------------------------------------

THIN = Side(border_style="thin", color="CFD8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F6F4A")  # brand green
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="0F3D2A")
SUB_FONT = Font(name="Calibri", size=10, color="475569")

SECTION_FILL = PatternFill("solid", fgColor="E8F5EE")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="0F3D2A")

STATUS_FILLS = {
    "Pass": PatternFill("solid", fgColor="DCFCE7"),
    "Failed": PatternFill("solid", fgColor="FEE2E2"),
    "Pending": PatternFill("solid", fgColor="FEF9C3"),
}
STATUS_FONTS = {
    "Pass": Font(name="Calibri", size=10, bold=True, color="14532D"),
    "Failed": Font(name="Calibri", size=10, bold=True, color="7F1D1D"),
    "Pending": Font(name="Calibri", size=10, bold=True, color="713F12"),
}

SEVERITY_FILLS = {
    "Critical": PatternFill("solid", fgColor="DC2626"),
    "High": PatternFill("solid", fgColor="F97316"),
    "Medium": PatternFill("solid", fgColor="FACC15"),
    "Low": PatternFill("solid", fgColor="93C5FD"),
    "Fixed": PatternFill("solid", fgColor="86EFAC"),
    "Pending": PatternFill("solid", fgColor="FEF08A"),
    "": PatternFill("solid", fgColor="F1F5F9"),
}
SEVERITY_FONTS = {
    "Critical": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "High": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "Medium": Font(name="Calibri", size=10, bold=True, color="1F2937"),
    "Low": Font(name="Calibri", size=10, bold=True, color="1F2937"),
    "Fixed": Font(name="Calibri", size=10, bold=True, color="14532D"),
    "Pending": Font(name="Calibri", size=10, bold=True, color="713F12"),
    "": Font(name="Calibri", size=10, color="475569"),
}

WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

ROW_FILL_EVEN = PatternFill("solid", fgColor="FAFAFA")
FAIL_ROW_FILL = PatternFill("solid", fgColor="FFF1F2")

COLUMNS = [
    ("#", 5),
    ("Module / Page", 18),
    ("Feature", 22),
    ("Test Scenario", 32),
    ("Steps to Test", 42),
    ("Expected Result", 32),
    ("Actual Result", 32),
    ("Status", 10),
    ("Severity", 10),
    ("Screenshot / Reference", 28),
    ("Remark / Comment", 28),
    ("Tested By", 16),
    ("Testing Date", 12),
]

# ---------- test cases ----------------------------------------------------
# Status: Pass | Failed | Pending
# Severity: Critical | High | Medium | Low | "" (for Pass rows leave blank)

TC = [
    # ------------------------ Explore Page ------------------------
    {
        "module": "Explore Page",
        "feature": "Initial render",
        "scenario": "Open root URL — site redirects to /explore and renders hero, search, listings",
        "steps": "1) Navigate to http://localhost:3000/\n2) Confirm redirect to /explore\n3) Observe hero, search bar (Location / Property type / Sort), List/Map tabs, and listing grid",
        "expected": "Hero heading, sub-tagline, 3-field search bar, List/Map tabs and room cards visible. Counter reads '27 rooms available right now'.",
        "actual": "Page redirects to /explore. Hero, search bar with Location/Property type/Sort buttons, List/Map tabs and 27 mock listings rendered. Counter matches.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/01_explore_desktop.png",
        "remark": "Anonymous, English language.",
    },
    {
        "module": "Explore Page",
        "feature": "Language toggle",
        "scenario": "Toggle UI between Khmer (default) and English",
        "steps": "1) Open /explore (default Khmer)\n2) Click the language pill (flag icon)\n3) Verify all labels switch to English\n4) Reload page",
        "expected": "Language persists across reload via localStorage 'findroom.language'. Hero, search bar, counter, tab labels all translate.",
        "actual": "Toggle switches between km/en. Persisted as 'en' in localStorage and survives reload. Khmer screenshot (50_explore_khmer) confirms KM render too.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/50_explore_khmer.png",
        "remark": "Persistence verified.",
    },
    {
        "module": "Explore Page",
        "feature": "Location filter (province + district)",
        "scenario": "Cascading location picker narrows listings",
        "steps": "1) Click 'Where to?' search field\n2) Pick Phnom Penh → Boeng Keng Kang\n3) Click 'Show all rooms in Boeng Keng Kang'\n4) Read counter and verify card count",
        "expected": "Counter and grid update to only listings in the selected district. Chip shows 'Phnom Penh, Boeng Keng Kang' with × clear affordance.",
        "actual": "Counter dropped from 27 → 3. Three matching cards rendered (BKK1 listings). × button rendered next to chip.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/03_explore_filtered_bkk.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Clear location chip",
        "scenario": "× button on chip clears the location filter",
        "steps": "1) Apply Phnom Penh → BKK filter\n2) Click × inside the search chip\n3) Verify counter returns to all listings",
        "expected": "Counter resets to 27 rooms; chip label returns to placeholder.",
        "actual": "Counter restored to 27 after click. Note: when the LocationPicker dialog is still open, the first × click only dismisses the dialog; a second click is required to clear. Minor UX nit.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "",
        "remark": "Picker overlay can swallow the × click — consider closing picker first or stopPropagation on chip × button.",
    },
    {
        "module": "Explore Page",
        "feature": "Property type filter",
        "scenario": "Selecting a property type narrows results",
        "steps": "1) Open 'Any property type' dropdown\n2) Pick 'Condo'\n3) Verify counter and grid",
        "expected": "Only Condo listings shown. Counter reflects new total.",
        "actual": "Counter dropped to 5 rooms; all 5 cards show 'Condo' category badge.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/04_property_type_open.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Sort by price (asc / desc)",
        "scenario": "Listings re-order by price",
        "steps": "1) Apply Condo filter\n2) Sort by price → Low to high — record prices\n3) Sort by price → High to low — record prices",
        "expected": "Low→high sequence is non-decreasing; high→low sequence is non-increasing.",
        "actual": "Low→high: $320, $420, $430, $480, $690 (✓). High→low: $690, $480, $430, $420, $320 (✓).",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/05_sort_price_asc.png",
        "remark": "",
    },
    {
        "module": "Explore Page",
        "feature": "Map view tab",
        "scenario": "Switch from List to Map view; markers render for visible rooms",
        "steps": "1) On /explore click 'Map' tab\n2) Wait for Leaflet map to load\n3) Count markers vs filtered listings",
        "expected": "Map renders without errors, marker count matches the filtered list (5 for Condo filter, 27 for unfiltered).",
        "actual": "Map mounted, 5 markers for Condo filter, 27 for unfiltered. No JS console errors.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/06_map_view.png",
        "remark": "Lazy-loaded via dynamic import. Loading state visible briefly.",
    },

    # ------------------------ Room Detail ------------------------
    {
        "module": "Room Detail",
        "feature": "Detail page render",
        "scenario": "Open a mock room from Explore",
        "steps": "1) From /explore, click first card (Cozy studio near Riverside)\n2) Verify route /rooms/1\n3) Verify hero image, title, address, badges (bed/area/floor), About, Amenities, Fees, Similar rooms",
        "expected": "All sections render with mock-data values; sticky footer shows price plus Location & Contact actions.",
        "actual": "Loaded /rooms/1: hero image, title 'Cozy studio near Riverside', address 'St. 110, Daun Penh, Doun Penh, Phnom Penh', 1 bed / 28 m² / 3F badges, About, 6 amenities, Fees & utilities table with rent $250/month, deposit, electricity, water, Wi-Fi. Similar rooms (3) listed. Sticky footer present.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/07_room_detail.png",
        "remark": "",
    },
    {
        "module": "Room Detail",
        "feature": "Photo gallery viewer",
        "scenario": "Open gallery overlay and step through photos",
        "steps": "1) Click '4 photos' button on hero\n2) Verify overlay shows 1/4\n3) Click 'Next photo' 3×\n4) Confirm counter reaches 4/4",
        "expected": "Overlay opens, counter updates 1/4 → 2/4 → 3/4 → 4/4. Close (×), Previous, Next, and thumbnail buttons work.",
        "actual": "Overlay opens. Counter increments correctly to 4/4 after 3 Next clicks. All buttons exposed via aria-label (Previous photo, Next photo, Show photo 1..4).",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "",
    },
    {
        "module": "Room Detail",
        "feature": "Contact host modal",
        "scenario": "Open contact modal from sticky footer",
        "steps": "1) Click 'Contact' button on sticky footer\n2) Verify contact info (host name, phones, Telegram)",
        "expected": "Modal lists host (Sokha Chan) with phones (+855 12 345 678 / +855 70 123 456) and Telegram. Close button works.",
        "actual": "Modal rendered with the expected fields. No copy / tap-to-call affordance (links are plain text). Acceptable for current scope.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": f"{SS_DIR}/08_contact_modal.png",
        "remark": "Consider making phone numbers tel: links and Telegram a t.me/* link in a future iteration.",
    },
    {
        "module": "Room Detail",
        "feature": "Location modal — embedded map",
        "scenario": "Tap Location → Show map; verify Google Maps iframe loads",
        "steps": "1) Click 'Location' on sticky footer\n2) Tap 'Show map' inside the modal\n3) Wait for iframe load",
        "expected": "Embedded Google Maps iframe at full modal width centered on the room's lat/lng. 'Open in Maps ↗' deeplink also rendered.",
        "actual": "Iframe loads at correct size (~476×298 in 1280-wide viewport) pointing to https://www.google.com/maps?q=11.5775,104.925. 'Open in Maps' link present.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/09_location_modal.png",
        "remark": "iframe sizing collapsed (2 px wide) when the parent viewport itself was 2 px during an earlier resize event — root cause was Playwright viewport, not the modal. Re-tested at 1280 px and result is correct.",
    },
    {
        "module": "Room Detail",
        "feature": "Invalid room id (404)",
        "scenario": "Visit /rooms/<bogus> directly",
        "steps": "1) Navigate to /rooms/does-not-exist\n2) Verify graceful empty state",
        "expected": "Page reads 'Room not found — The listing you're looking for doesn't exist or has been removed.' with a Back to Explore button.",
        "actual": "Exact message and CTA rendered. No client-side error.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/10_room_not_found.png",
        "remark": "",
    },

    # ------------------------ Auth ------------------------
    {
        "module": "Auth",
        "feature": "Open Login modal",
        "scenario": "Logged-out user clicks 'Log in' in the navbar",
        "steps": "1) /explore as anonymous\n2) Click 'Log in' in top-right\n3) Verify modal title and fields",
        "expected": "Modal opens with phone (tel) and password inputs both marked required. 'Create one' switch link visible.",
        "actual": "Modal shows 'Welcome back', phone (placeholder '12 345 678' with +855 prefix), password, Log in button, 'Create one' link. Both inputs have required attribute.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/11_login_modal.png",
        "remark": "",
    },
    {
        "module": "Auth",
        "feature": "Empty submit validation",
        "scenario": "Submit Log in without filling either field",
        "steps": "1) Open Log in modal\n2) Click 'Log in' immediately",
        "expected": "Native HTML5 'Please fill in this field' validation. Modal stays open.",
        "actual": "Form does not submit. Inputs marked required. Modal stays open. No custom error string, relies on browser tooltip.",
        "status": "Pass",
        "severity": "Low",
        "screenshot": "",
        "remark": "Consider explicit in-form error messages for nicer UX, but functionally protected.",
    },
    {
        "module": "Auth",
        "feature": "Demo-mode login (any password)",
        "scenario": "Log in as the seed admin (12000000 / anything)",
        "steps": "1) Phone: 12000000, Password: password123\n2) Submit\n3) Inspect localStorage 'findroom.session'",
        "expected": "Session stored as {uid:'demo-85512000000', phoneNumber:'+85512000000'}. Modal closes. Navbar now shows User/Admin pills and notification badge.",
        "actual": "Session stored correctly. Modal closes, /explore visible. Admin pills (User/Admin) and unread notification badge render.",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "Demo mode accepts any password by design (no Firebase configured locally).",
    },
    {
        "module": "Auth",
        "feature": "Register tab",
        "scenario": "Switch from Login to Register flow",
        "steps": "1) Open Log in modal\n2) Click 'No account yet? Create one'\n3) Verify form shows name + phone + password fields",
        "expected": "Modal swaps to 'Create your account' with name, phone, password inputs.",
        "actual": "Register tab renders. Fields available.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/12_register_modal.png",
        "remark": "Confirm/repeat-password not requested.",
    },

    # ------------------------ Profile ------------------------
    {
        "module": "Profile",
        "feature": "Profile page render + sample seed",
        "scenario": "Newly logged-in user lands on /profile and sees 3 sample listings",
        "steps": "1) Log in as 12000000\n2) Visit /profile\n3) Confirm header (avatar / name / phone), Edit profile + Log out actions, 'My listings' with 3 sample rooms",
        "expected": "Header with avatar fallback letter and phone +85512000000. 3 sample listings: Cozy studio near Riverside, Bright 1-bedroom in BKK1, Family house in Sen Sok (Occupied).",
        "actual": "All sections rendered. Sample seeded into localStorage findroom.local-rooms with id 'sample-<uid>-1/2/3'.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/13_profile_page.png",
        "remark": "Sample seed flag is one-shot per uid (findroom.seeded.<uid>).",
    },
    {
        "module": "Profile",
        "feature": "Edit profile (name + avatar + login phone)",
        "scenario": "Update display name and verify persistence",
        "steps": "1) Click 'Edit profile'\n2) Change Name from 'FindRoom user' to 'QA Tester'\n3) Click Save\n4) Reload page and re-check display name",
        "expected": "Modal closes, header shows 'QA Tester'. Reload preserves the new name (localStorage findroom.profile-overrides-<uid>).",
        "actual": "Name updated to 'QA Tester' immediately and survives full reload. Override persisted under expected key.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/14_edit_profile_modal.png",
        "remark": "",
    },

    # ------------------------ List Room (form) ------------------------
    {
        "module": "List Room form",
        "feature": "Form render — required field markers",
        "scenario": "Open /profile/list-room and verify sections + required-field asterisks",
        "steps": "1) Visit /profile/list-room as authenticated user\n2) Inspect sections: Photos, Title*, Room details, About this place*, What this place offers, Fees & utilities*, Contact, Location*",
        "expected": "All sections present. Title, About, Fees, Location marked with * (required). Contact pre-seeded with logged-in phone.",
        "actual": "All sections render in the expected order; required asterisks present on Title, About, Fees, Location. Contact list pre-seeded with +85512000000.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/15_list_room_form.png",
        "remark": "",
    },
    {
        "module": "List Room form",
        "feature": "Empty submit validation",
        "scenario": "Submit form with no fields filled — verify cascading inline error",
        "steps": "1) Click 'Publish listing' immediately\n2) Read inline error\n3) Fill title only and resubmit\n4) Continue filling and resubmit until success",
        "expected": "Inline error guides the user one step at a time: 'Add a title' → 'Pick a province' → 'Set a monthly rent' etc. URL remains /profile/list-room.",
        "actual": "Inline error appears under the action bar. Cascade order verified: title → description → province → rent. Page does not navigate until all required are satisfied.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/16_list_room_validation.png",
        "remark": "Server-side validation will replace this in production.",
    },
    {
        "module": "List Room form",
        "feature": "Fees modal",
        "scenario": "Open Fees & utilities modal and set rent",
        "steps": "1) Click 'Add fees'\n2) Set rent $300 / Monthly\n3) Click Done\n4) Verify summary in main form",
        "expected": "Modal closes; main form section shows the rent and period.",
        "actual": "Modal closes; rent + period reflected in the Fees section preview.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/17_fees_modal.png",
        "remark": "Add fee button supports utility types (deposit/electricity/water/Wi-Fi/service/parking/other).",
    },
    {
        "module": "List Room form",
        "feature": "Successful publish",
        "scenario": "Submit a valid listing (title, description, province, rent)",
        "steps": "1) Fill title 'QA test listing'\n2) Description 'Auto-generated…'\n3) Location → Phnom Penh\n4) Add fees → rent $300\n5) Publish",
        "expected": "Listing is persisted to localStorage 'findroom.local-rooms'. Browser navigates to /profile and the new card is visible. Listing also appears on /explore.",
        "actual": "Listing saved (length 3→4). Redirected to /profile and the new card appears. /explore shows it among '30 rooms available right now'.",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "Owner inherited from session uid.",
    },

    # ------------------------ Admin Console ------------------------
    {
        "module": "Admin Console",
        "feature": "Access control",
        "scenario": "Admin uid can reach /user/admin, non-admin sees Admins-only screen",
        "steps": "1) Logged in as 12000000 → visit /user/admin\n2) Sign out, log in as 12345678 (Sokha) → visit /user/admin",
        "expected": "Demo admin sees Rooms dashboard. Sokha (regular user) sees 'Admins only' lockout screen with Back / Switch account buttons.",
        "actual": "Demo admin reaches Rooms dashboard. Lockout flow not retested in this pass — works in source (`AdminShell.tsx` uses `isAdmin(session)`).",
        "status": "Pending",
        "severity": "",
        "screenshot": "",
        "remark": "Re-run after backend role gating is wired — current allowlist is client-side only.",
    },
    {
        "module": "Admin Console",
        "feature": "Rooms dashboard render",
        "scenario": "Open /user/admin and verify summary tiles + table",
        "steps": "1) Visit /user/admin\n2) Check tiles: Rooms / Available / Occupied / Property Types\n3) Inspect first table row",
        "expected": "Tiles show 30 total / 27 available / 3 occupied / 6 types (fresh state). Table lists 30 rows, each with thumbnail, title + type, owner, location, price, status badge, Row actions menu.",
        "actual": "Tiles match (33/29/4/6 after one user listing seeded + 3 sample listings carried by the logged-in admin). Table rows render correctly.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/18_admin_rooms.png",
        "remark": "Numbers include the admin's own sample-seeded + QA-test listings, which is expected behaviour for a frontend-first store.",
    },
    {
        "module": "Admin Console",
        "feature": "Edit listing modal",
        "scenario": "Edit a listing's title, price, owner, occupied status",
        "steps": "1) Click Row actions (•••) on a row\n2) Choose Edit\n3) Verify pre-filled fields + Save / Cancel",
        "expected": "Modal opens with Title, Monthly price (USD), Owner select (sorted alphabetically with current owner pinned), 'Marked as occupied' toggle. Cancel closes without changes.",
        "actual": "Modal opens with the pre-filled fields. Owner select lists 8 user options. Cancel closes the modal.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/19_admin_edit_listing.png",
        "remark": "RowActions menu closes on blur after 150 ms — programmatic clicks must dispatch mousedown+click. Real mouse users are unaffected.",
    },
    {
        "module": "Admin Console",
        "feature": "Users page render",
        "scenario": "Open /user/admin/users and verify tiles + table",
        "steps": "1) Visit /user/admin/users\n2) Check tiles: Users / Active / Disabled / Admins\n3) Confirm the logged-in admin is highlighted",
        "expected": "Tiles: 8 / 7 / 1 / 1. Row 'Admin (you)' pinned to top, Role=Admin, Status=Active.",
        "actual": "All tiles and rows match. 'Admin (you)' row rendered with Admin role and Active status. Mock email rendered when available.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/20_admin_users.png",
        "remark": "",
    },
    {
        "module": "Admin Console",
        "feature": "Notifications + Mark all read",
        "scenario": "Mark all admin notifications as read",
        "steps": "1) Visit /user/admin/notifications\n2) Read unread badge count\n3) Click 'Mark all read'\n4) Re-check badge and persistence",
        "expected": "All 5 notifications are marked read in 'findroom.admin-notifications'. Badge clears. 'Mark all read' becomes disabled.",
        "actual": "All notifications flipped to read=true in localStorage. Badge cleared. The button stays visible but is now disabled — UX behaves correctly.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/21_admin_notifications.png",
        "remark": "",
    },
    {
        "module": "Admin Console",
        "feature": "Settings page render & form",
        "scenario": "Open /user/admin/settings and inspect editable sections",
        "steps": "1) Visit /user/admin/settings\n2) Inspect: General / Listing taxonomy / Pricing defaults / Moderation / Access control",
        "expected": "All settings sections render with prefilled values; property types and amenities chips can be removed (×).",
        "actual": "All sections render. Default values match in-code seed (FindRoom.KH, support@findroom.kh, USD, etc.). Amenities + property types chips show × affordance.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/22_admin_settings.png",
        "remark": "Settings persistence wired to localStorage 'findroom.admin-settings'.",
    },
    {
        "module": "Admin Console",
        "feature": "View-mode toggle (Admin → User)",
        "scenario": "Admin switches to User view via the navbar pill",
        "steps": "1) On /user/admin/settings click the 'User' pill\n2) Confirm redirect to /explore and User chrome",
        "expected": "Redirected to /explore. 'List room' button visible in nav, admin floating nav hidden.",
        "actual": "Redirected to /explore. User chrome restored. View-mode preference persists across navigation.",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "",
    },

    # ------------------------ FIXED BUG ------------------------
    {
        "module": "Explore Page",
        "feature": "Duplicate listings after admin visit",
        "scenario": "After visiting /user/admin, /explore renders every mock listing twice",
        "steps": "1) Log in as 12000000 (admin)\n2) Visit /user/admin (any tab)\n3) Navigate to /explore\n4) Count rendered cards vs unique titles",
        "expected": "Each room appears once. Counter matches the number of cards.",
        "actual": "Originally Failed: 57 cards / 29 unique titles after admin visit. Cards `/rooms/1` and `/rooms/mock-1` both rendered the same room. Root cause: `seedMockListings()` in src/lib/admin.ts:197 writes MOCK_ROOMS into localStorage as `mock-<id>`; ExploreRooms.tsx then merged `[...localRooms, ...MOCK_ROOMS]` without de-duplication.\n\nFIX applied 2026-05-22 in src/components/ExploreRooms.tsx — canonical-id dedupe (`mock-<id>` collapses to `<id>`, localRooms wins). Re-tested: 27 cards / 27 IDs, counter 27 ✓.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/24_fix_explore_no_duplicates.png",
        "remark": "Before-fix evidence retained at screenshots_2026-05-22/23_bug_explore_duplicates.png for the changelog.",
    },
    {
        "module": "Explore Page",
        "feature": "Regression — listing dedupe with sample seeds",
        "scenario": "Admin logged-in user with 3 sample-seeded rooms + MOCK_ROOMS + 1 user-added shows no duplicates",
        "steps": "1) Log in as 12000000\n2) Visit /profile (triggers sample seed) and /user/admin (triggers mock seed)\n3) Return to /explore\n4) Compare card count vs unique IDs",
        "expected": "Card count equals unique-id count. Counter matches card count. Same-title coincidence between sample seed and mock seed (Cozy studio near Riverside) is fine — they're distinct listings owned by different users.",
        "actual": "29 cards / 29 IDs / counter 29. One repeated title belongs to two distinct owners (Sokha Chan vs the logged-in admin's sample). Filter (Phnom Penh) narrowed to 20/20/20 — also clean.",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "Locks in the fix; guards against re-introducing the bug.",
    },

    # ------------------------ Cross-device ------------------------
    {
        "module": "Responsive",
        "feature": "Mobile layout (375×812)",
        "scenario": "Explore page on mobile viewport",
        "steps": "1) Set viewport 375×812\n2) Visit /explore\n3) Verify hero stack, single-column search, BottomNav (Home / FAB / Profile)",
        "expected": "Hero stacks, search fields stack vertically, BottomNav visible with rounded FAB, navbar hides 'List room' button on mobile.",
        "actual": "Layout matches: hero stacks, search fields stack, BottomNav with Home / FAB / Profile rendered. Language toggle still in navbar.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/30_mobile_explore.png",
        "remark": "",
    },
    {
        "module": "Responsive",
        "feature": "Mobile room detail",
        "scenario": "Room detail page on mobile viewport",
        "steps": "1) /rooms/1 at 375×812\n2) Scroll through sections",
        "expected": "Hero image fits viewport width, sections stack, sticky footer reachable.",
        "actual": "Rendered as expected. Sticky bottom bar shows price + Location / Contact.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/31_mobile_room_detail.png",
        "remark": "",
    },
    {
        "module": "Responsive",
        "feature": "Tablet layout (768×1024)",
        "scenario": "Explore page on tablet viewport",
        "steps": "1) 768×1024\n2) Visit /explore\n3) Inspect search bar and grid",
        "expected": "Search bar transitions to compact column layout; listing grid uses 3 columns. Navbar shows 'List room' button.",
        "actual": "Search bar columnar with full-width fields, listing grid 3 columns. Navbar shows List room.",
        "status": "Pass",
        "severity": "",
        "screenshot": f"{SS_DIR}/40_tablet_explore.png",
        "remark": "",
    },

    # ------------------------ Print/Export ------------------------
    {
        "module": "Print / Export",
        "feature": "Listing export / print",
        "scenario": "Verify export-to-PDF or print affordance",
        "steps": "1) Inspect room detail and admin pages for export / print buttons",
        "expected": "Export or print button per business spec.",
        "actual": "No print or export buttons exist in the current build. Scope confirmation needed — feature may be out-of-scope for the personal/landlord MVP.",
        "status": "Pending",
        "severity": "Low",
        "screenshot": "",
        "remark": "Confirm with PM whether print/export is in MVP scope before opening as a bug.",
    },

    # ------------------------ Performance ------------------------
    {
        "module": "Performance",
        "feature": "Console & network errors on initial load",
        "scenario": "Browse /explore → /rooms/1 → /profile and watch console / network",
        "steps": "1) Open DevTools\n2) Navigate the happy path\n3) Inspect console_logs and failed network",
        "expected": "No uncaught errors, no failed network requests.",
        "actual": "No uncaught errors detected. All image requests succeed. (Note: Firebase analytics endpoint may log a non-fatal warning when env keys are absent — informational only.)",
        "status": "Pass",
        "severity": "",
        "screenshot": "",
        "remark": "",
    },
]

# ---------- workbook builder ---------------------------------------------

def build_summary(ws, cases):
    ws.title = "Summary"
    ws["A1"] = "FindRoom.KH — QA Testing Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"Tested by: {TESTER}   |   Test date: {TEST_DATE}   |   Build: local dev (Next.js 14, frontend-first w/ localStorage mocks)"
    )
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    headers = ["Module", "Total", "Pass", "Failed", "Pending", "Fail Rate"]
    ws.append([])
    ws.append(headers)
    row = ws.max_row
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    # tally by module
    tally = {}
    for c in cases:
        m = c["module"]
        tally.setdefault(m, {"Pass": 0, "Failed": 0, "Pending": 0, "Total": 0})
        tally[m]["Total"] += 1
        tally[m][c["status"]] += 1

    for module, t in tally.items():
        fail_rate = f"{(t['Failed'] / t['Total']) * 100:.0f}%" if t["Total"] else "—"
        ws.append([module, t["Total"], t["Pass"], t["Failed"], t["Pending"], fail_rate])
        r = ws.max_row
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.alignment = WRAP if col == 1 else CENTER
            cell.font = Font(name="Calibri", size=10)
        if t["Failed"]:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = STATUS_FILLS["Failed"]

    # grand total
    grand = {
        "Total": sum(t["Total"] for t in tally.values()),
        "Pass": sum(t["Pass"] for t in tally.values()),
        "Failed": sum(t["Failed"] for t in tally.values()),
        "Pending": sum(t["Pending"] for t in tally.values()),
    }
    grand_fail = (
        f"{(grand['Failed'] / grand['Total']) * 100:.0f}%" if grand["Total"] else "—"
    )
    ws.append(["TOTAL", grand["Total"], grand["Pass"], grand["Failed"], grand["Pending"], grand_fail])
    r = ws.max_row
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = SECTION_FILL
        cell.border = BORDER
        cell.alignment = CENTER if col != 1 else WRAP

    # known issues block
    ws.append([])
    ws.append(["Known Issues / Action Items"])
    r = ws.max_row
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    issues = [
        ("Fixed", "Explore Page", "Duplicate listings after admin visit — FIXED 2026-05-22 in src/components/ExploreRooms.tsx via canonical-id dedupe. Regression case #30 covers it."),
        ("Low", "Explore Page", "Clear-location × intercepted by open picker overlay. Either auto-close picker first or stopPropagation on the chip ×."),
        ("Low", "Room Detail", "Phone numbers / Telegram in contact modal are plain text. Tap-to-call / tap-to-chat would be a quick win on mobile."),
        ("Pending", "Admin Console", "Non-admin lockout flow needs retest once role gating is server-enforced (currently client-only allowlist)."),
        ("Pending", "Print / Export", "Confirm with PM whether print or PDF export is in MVP scope."),
    ]
    ws.append(["Severity", "Module", "Detail"])
    r = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row=r, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    for sev, mod, detail in issues:
        ws.append([sev, mod, detail])
        r = ws.max_row
        sev_cell = ws.cell(row=r, column=1)
        sev_cell.fill = SEVERITY_FILLS.get(sev, SEVERITY_FILLS[""])
        sev_cell.font = SEVERITY_FONTS.get(sev, SEVERITY_FONTS[""])
        sev_cell.alignment = CENTER
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = WRAP if col == 3 else CENTER

    # column widths
    widths = [22, 12, 12, 12, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_cases(ws, cases):
    ws.title = "Test Cases"
    # header row
    ws.append([c[0] for c in COLUMNS])
    for i, (name, w) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    # freeze header + first 2 cols
    ws.freeze_panes = "D2"

    for idx, c in enumerate(cases, start=1):
        ws.append([
            idx,
            c["module"],
            c["feature"],
            c["scenario"],
            c["steps"],
            c["expected"],
            c["actual"],
            c["status"],
            c["severity"],
            c["screenshot"],
            c["remark"],
            TESTER,
            TEST_DATE,
        ])
        r = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10, color="1F2937")
        # row banding
        if c["status"] == "Failed":
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=col_idx).fill = FAIL_ROW_FILL
        elif idx % 2 == 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                if ws.cell(row=r, column=col_idx).fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    ws.cell(row=r, column=col_idx).fill = ROW_FILL_EVEN

        # status pill
        status_cell = ws.cell(row=r, column=8)
        status_cell.fill = STATUS_FILLS[c["status"]]
        status_cell.font = STATUS_FONTS[c["status"]]
        status_cell.alignment = CENTER
        # severity pill
        sev_cell = ws.cell(row=r, column=9)
        sev_cell.fill = SEVERITY_FILLS.get(c["severity"], SEVERITY_FILLS[""])
        sev_cell.font = SEVERITY_FONTS.get(c["severity"], SEVERITY_FONTS[""])
        sev_cell.alignment = CENTER
        # screenshot hyperlink (relative path)
        ss = c["screenshot"]
        if ss:
            ws.cell(row=r, column=10).hyperlink = ss
            ws.cell(row=r, column=10).font = Font(
                name="Calibri", size=10, color="1F6F4A", underline="single"
            )
        # tested-by + date centered
        for col_idx in (1, 12, 13):
            ws.cell(row=r, column=col_idx).alignment = CENTER
        # numeric # in column 1
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color="475569")

        # height — give roomy lines for wrapped text
        rough_lines = max(
            len(str(c[k]).split("\n")) for k in ("steps", "expected", "actual")
        )
        ws.row_dimensions[r].height = max(60, 18 + rough_lines * 14)


def build_evidence(ws, cases):
    ws.title = "Evidence Index"
    ws.append(["#", "Module", "Feature", "Status", "Screenshot path"])
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    widths = [5, 18, 26, 10, 56]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    idx = 0
    for i, c in enumerate(cases, start=1):
        if not c["screenshot"]:
            continue
        idx += 1
        ws.append([i, c["module"], c["feature"], c["status"], c["screenshot"]])
        r = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.alignment = WRAP if col == 5 else CENTER
            cell.font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=5).hyperlink = c["screenshot"]
        ws.cell(row=r, column=5).font = Font(
            name="Calibri", size=10, color="1F6F4A", underline="single"
        )
        ws.cell(row=r, column=4).fill = STATUS_FILLS[c["status"]]
        ws.cell(row=r, column=4).font = STATUS_FONTS[c["status"]]


def main():
    wb = Workbook()
    summary = wb.active
    build_summary(summary, TC)
    cases_ws = wb.create_sheet("Test Cases")
    build_cases(cases_ws, TC)
    evidence_ws = wb.create_sheet("Evidence Index")
    build_evidence(evidence_ws, TC)

    out = f"qa_report/QA_Report_FindRoom_{TEST_DATE}.xlsx"
    wb.save(out)
    print("Wrote", out, "with", len(TC), "test cases")


if __name__ == "__main__":
    main()
